"""Kait Sidekick -- File Processing Engine

Extracts content from documents, images, code, audio, video, and archives
to provide as context for LLM interactions.  Follows the strategy pattern
(like ToolRegistry in tools.py).

All handlers check ``is_available()`` at runtime via try-import so the
system degrades gracefully when optional packages are missing.

No Qt dependencies -- this module is pure Python + stdlib + optional libs.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import mimetypes
import os
import re
import tarfile
import time
import wave
import zipfile
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

log = logging.getLogger("kait.file_processor")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
MAX_CONTENT_LENGTH = 100_000  # characters
MAX_ATTACHMENTS = 10

# Default allowed roots (mirrors tools.py sandbox policy)
_DEFAULT_ALLOWED_ROOTS: List[Path] = [
    Path.home() / ".kait",
    Path.home() / "Documents",
    Path.home() / "Desktop",
    Path.home() / "Downloads",
]


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class FileAttachment:
    """Tracks an attached file through the UI pipeline."""

    path: str
    name: str
    size: int
    category: str = "unknown"
    status: str = "pending"  # pending | processing | ready | error
    result: Optional["FileProcessingResult"] = None
    progress: int = 0


@dataclass
class FileProcessingResult:
    """Result of processing a single attached file."""

    file_path: str
    file_name: str
    file_size: int
    mime_type: str
    category: str
    content: str  # extracted text
    metadata: Dict[str, Any] = field(default_factory=dict)
    thumbnail: Optional[bytes] = None
    success: bool = True
    error: Optional[str] = None
    processing_time_ms: int = 0


# ---------------------------------------------------------------------------
# Handler ABC (strategy pattern)
# ---------------------------------------------------------------------------

class FileTypeHandler(ABC):
    """Base class for file-type handlers."""

    @property
    @abstractmethod
    def extensions(self) -> List[str]:
        """Supported file extensions (lowercase, with leading dot)."""
        ...

    @property
    @abstractmethod
    def category(self) -> str:
        """Category label: document, code, image, video, audio, archive."""
        ...

    @property
    def requires(self) -> List[str]:
        """Optional package names needed for full functionality."""
        return []

    @abstractmethod
    def process(
        self,
        path: Path,
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> FileProcessingResult:
        ...

    def is_available(self) -> bool:
        """Return True if this handler can operate (all deps present)."""
        return True


# ===========================================================================
# Handlers
# ===========================================================================

class PlainTextHandler(FileTypeHandler):
    """Handles plain text files: .txt, .md, .csv, .json, .yaml, .yml, .log, .ini, .cfg, .toml"""

    @property
    def extensions(self) -> List[str]:
        return [
            ".txt", ".md", ".csv", ".json", ".yaml", ".yml",
            ".log", ".ini", ".cfg", ".toml",
        ]

    @property
    def category(self) -> str:
        return "document"

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        mime = mimetypes.guess_type(str(path))[0] or "text/plain"
        metadata: Dict[str, Any] = {"encoding": "utf-8"}

        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=path.stat().st_size,
                mime_type=mime, category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        line_count = text.count("\n") + 1
        metadata["line_count"] = line_count
        metadata["char_count"] = len(text)

        # Extra metadata for structured formats
        ext = path.suffix.lower()
        if ext == ".csv":
            try:
                reader = csv.reader(io.StringIO(text))
                rows = list(reader)
                metadata["row_count"] = len(rows)
                metadata["column_count"] = len(rows[0]) if rows else 0
            except Exception:
                pass
        elif ext == ".json":
            try:
                parsed = json.loads(text)
                metadata["json_type"] = type(parsed).__name__
                if isinstance(parsed, list):
                    metadata["item_count"] = len(parsed)
                elif isinstance(parsed, dict):
                    metadata["key_count"] = len(parsed)
            except Exception:
                pass

        if progress_callback:
            progress_callback(100)

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=path.stat().st_size,
            mime_type=mime, category=self.category,
            content=text[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class CodeFileHandler(FileTypeHandler):
    """Handles source-code files with language metadata."""

    _LANG_MAP = {
        ".py": "python", ".js": "javascript", ".ts": "typescript",
        ".html": "html", ".css": "css", ".sql": "sql", ".sh": "shell",
        ".go": "go", ".rs": "rust", ".java": "java",
        ".c": "c", ".cpp": "cpp", ".h": "c-header",
        ".rb": "ruby", ".php": "php", ".swift": "swift",
        ".kt": "kotlin", ".r": "r", ".lua": "lua",
    }

    @property
    def extensions(self) -> List[str]:
        return list(self._LANG_MAP.keys())

    @property
    def category(self) -> str:
        return "code"

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        mime = mimetypes.guess_type(str(path))[0] or "text/plain"
        lang = self._LANG_MAP.get(path.suffix.lower(), "unknown")

        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=path.stat().st_size,
                mime_type=mime, category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        line_count = text.count("\n") + 1
        metadata: Dict[str, Any] = {
            "language": lang,
            "line_count": line_count,
            "char_count": len(text),
        }

        # Count functions/classes for Python
        if lang == "python":
            metadata["def_count"] = len(re.findall(r"^\s*def\s+", text, re.MULTILINE))
            metadata["class_count"] = len(re.findall(r"^\s*class\s+", text, re.MULTILINE))

        if progress_callback:
            progress_callback(100)

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=path.stat().st_size,
            mime_type=mime, category=self.category,
            content=text[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class PDFHandler(FileTypeHandler):
    """Handles PDF files via PyPDF2 (optional)."""

    @property
    def extensions(self) -> List[str]:
        return [".pdf"]

    @property
    def category(self) -> str:
        return "document"

    @property
    def requires(self) -> List[str]:
        return ["PyPDF2"]

    def is_available(self) -> bool:
        try:
            import PyPDF2  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size
        metadata: Dict[str, Any] = {}

        if not self.is_available():
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/pdf", category=self.category,
                content=f"[PDF file: {path.name} ({size:,} bytes) -- install PyPDF2 for text extraction]",
                metadata={"pages": "unknown"}, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        try:
            import PyPDF2
            pages_text: List[str] = []
            with open(path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                metadata["page_count"] = len(reader.pages)
                info = reader.metadata
                if info:
                    for key in ("title", "author", "subject", "creator"):
                        val = getattr(info, key, None)
                        if val:
                            metadata[key] = str(val)
                total = len(reader.pages)
                for i, page in enumerate(reader.pages):
                    pages_text.append(page.extract_text() or "")
                    if progress_callback:
                        progress_callback(int((i + 1) / total * 100))
            content = "\n\n".join(pages_text)
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/pdf", category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type="application/pdf", category=self.category,
            content=content[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class DocxHandler(FileTypeHandler):
    """Handles .docx files via python-docx (optional)."""

    @property
    def extensions(self) -> List[str]:
        return [".docx"]

    @property
    def category(self) -> str:
        return "document"

    @property
    def requires(self) -> List[str]:
        return ["python-docx"]

    def is_available(self) -> bool:
        try:
            import docx  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size

        if not self.is_available():
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                category=self.category,
                content=f"[DOCX file: {path.name} ({size:,} bytes) -- install python-docx for text extraction]",
                metadata={}, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        try:
            import docx
            doc = docx.Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            metadata: Dict[str, Any] = {"paragraph_count": len(paragraphs)}
            core = doc.core_properties
            if core.title:
                metadata["title"] = core.title
            if core.author:
                metadata["author"] = core.author
            content = "\n\n".join(paragraphs)
            if progress_callback:
                progress_callback(100)
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            category=self.category, content=content[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class SpreadsheetHandler(FileTypeHandler):
    """Handles .xlsx/.xls via openpyxl (optional), .csv via stdlib."""

    @property
    def extensions(self) -> List[str]:
        return [".xlsx", ".xls"]

    @property
    def category(self) -> str:
        return "document"

    @property
    def requires(self) -> List[str]:
        return ["openpyxl"]

    def is_available(self) -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size

        if not self.is_available():
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                category=self.category,
                content=f"[Spreadsheet: {path.name} ({size:,} bytes) -- install openpyxl for content extraction]",
                metadata={}, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheets_data: List[str] = []
            metadata: Dict[str, Any] = {"sheet_count": len(wb.sheetnames)}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows: List[str] = [f"## Sheet: {sheet_name}"]
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    rows.append("\t".join(cells))
                    row_count += 1
                    if row_count >= 500:  # cap rows per sheet
                        rows.append(f"... [truncated after {row_count} rows]")
                        break
                sheets_data.append("\n".join(rows))
                metadata[f"sheet_{sheet_name}_rows"] = row_count
            wb.close()

            content = "\n\n".join(sheets_data)
            if progress_callback:
                progress_callback(100)
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            category=self.category, content=content[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class ImageHandler(FileTypeHandler):
    """Handles images via Pillow (already in sidekick deps)."""

    @property
    def extensions(self) -> List[str]:
        return [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff"]

    @property
    def category(self) -> str:
        return "image"

    @property
    def requires(self) -> List[str]:
        return ["Pillow"]

    def is_available(self) -> bool:
        try:
            from PIL import Image  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size
        mime = mimetypes.guess_type(str(path))[0] or "image/unknown"

        if not self.is_available():
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category,
                content=f"[Image: {path.name} ({size:,} bytes)]",
                metadata={}, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        try:
            from PIL import Image
            from PIL.ExifTags import TAGS

            img = Image.open(path)
            metadata: Dict[str, Any] = {
                "width": img.width,
                "height": img.height,
                "format": img.format or path.suffix.upper(),
                "mode": img.mode,
            }

            # Extract EXIF
            exif_data = {}
            try:
                raw_exif = img.getexif()
                if raw_exif:
                    for tag_id, val in raw_exif.items():
                        tag_name = TAGS.get(tag_id, tag_id)
                        if isinstance(val, (str, int, float)):
                            exif_data[str(tag_name)] = val
                    if exif_data:
                        metadata["exif"] = exif_data
            except Exception:
                pass

            content = (
                f"[Image: {path.name}]\n"
                f"Dimensions: {img.width}x{img.height}\n"
                f"Format: {img.format or path.suffix}\n"
                f"Mode: {img.mode}"
            )
            if exif_data:
                content += f"\nEXIF: {json.dumps(exif_data, default=str)}"

            # Generate thumbnail bytes for UI preview
            thumbnail = None
            try:
                thumb = img.copy()
                thumb.thumbnail((128, 128))
                buf = io.BytesIO()
                thumb.save(buf, format="PNG")
                thumbnail = buf.getvalue()
            except Exception:
                pass

            img.close()
            if progress_callback:
                progress_callback(100)
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type=mime, category=self.category,
            content=content[:MAX_CONTENT_LENGTH],
            metadata=metadata, thumbnail=thumbnail, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class ImageOCRHandler(FileTypeHandler):
    """Extends ImageHandler with OCR via pytesseract (optional)."""

    @property
    def extensions(self) -> List[str]:
        return [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff"]

    @property
    def category(self) -> str:
        return "image"

    @property
    def requires(self) -> List[str]:
        return ["pytesseract", "Pillow"]

    def is_available(self) -> bool:
        try:
            import pytesseract  # noqa: F401
            from PIL import Image  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()

        # Fall back to ImageHandler if pytesseract not available
        if not self.is_available():
            return ImageHandler().process(path, progress_callback)

        try:
            import pytesseract
            from PIL import Image

            base_result = ImageHandler().process(path, progress_callback=None)

            img = Image.open(path)
            ocr_text = pytesseract.image_to_string(img).strip()
            img.close()

            if ocr_text:
                base_result.content += f"\n\nOCR Text:\n{ocr_text}"
                base_result.metadata["ocr_chars"] = len(ocr_text)

            base_result.content = base_result.content[:MAX_CONTENT_LENGTH]
            base_result.processing_time_ms = int((time.monotonic() - start) * 1000)
            if progress_callback:
                progress_callback(100)
            return base_result
        except Exception:
            # Fallback to plain image handling
            return ImageHandler().process(path, progress_callback)


class VideoHandler(FileTypeHandler):
    """Handles video files via opencv-python-headless (optional)."""

    @property
    def extensions(self) -> List[str]:
        return [".mp4", ".avi", ".mov", ".mkv", ".webm"]

    @property
    def category(self) -> str:
        return "video"

    @property
    def requires(self) -> List[str]:
        return ["opencv-python-headless"]

    def is_available(self) -> bool:
        try:
            import cv2  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size
        mime = mimetypes.guess_type(str(path))[0] or "video/unknown"

        if not self.is_available():
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category,
                content=f"[Video: {path.name} ({size:,} bytes) -- install opencv-python-headless for metadata extraction]",
                metadata={}, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        try:
            import cv2
            cap = cv2.VideoCapture(str(path))
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            duration = frame_count / fps if fps > 0 else 0

            metadata: Dict[str, Any] = {
                "width": width,
                "height": height,
                "fps": round(fps, 2),
                "frame_count": frame_count,
                "duration_seconds": round(duration, 2),
            }

            content = (
                f"[Video: {path.name}]\n"
                f"Resolution: {width}x{height}\n"
                f"Duration: {duration:.1f}s\n"
                f"FPS: {fps:.1f}\n"
                f"Frames: {frame_count}"
            )

            cap.release()
            if progress_callback:
                progress_callback(100)
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type=mime, category=self.category,
            content=content[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class AudioHandler(FileTypeHandler):
    """Handles audio files: .wav via stdlib, others via pydub (optional)."""

    @property
    def extensions(self) -> List[str]:
        return [".mp3", ".wav", ".ogg", ".flac"]

    @property
    def category(self) -> str:
        return "audio"

    @property
    def requires(self) -> List[str]:
        return ["pydub"]

    def is_available(self) -> bool:
        """Always available -- .wav handled by stdlib, others metadata only."""
        return True

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size
        mime = mimetypes.guess_type(str(path))[0] or "audio/unknown"
        metadata: Dict[str, Any] = {}

        ext = path.suffix.lower()

        # .wav via stdlib
        if ext == ".wav":
            try:
                with wave.open(str(path), "rb") as wf:
                    channels = wf.getnchannels()
                    sample_width = wf.getsampwidth()
                    framerate = wf.getframerate()
                    frames = wf.getnframes()
                    duration = frames / framerate if framerate > 0 else 0
                    metadata = {
                        "channels": channels,
                        "sample_width": sample_width,
                        "framerate": framerate,
                        "duration_seconds": round(duration, 2),
                    }
                    content = (
                        f"[Audio: {path.name}]\n"
                        f"Format: WAV\n"
                        f"Duration: {duration:.1f}s\n"
                        f"Channels: {channels}\n"
                        f"Sample rate: {framerate}Hz"
                    )
                if progress_callback:
                    progress_callback(100)
                return FileProcessingResult(
                    file_path=str(path), file_name=path.name, file_size=size,
                    mime_type=mime, category=self.category,
                    content=content, metadata=metadata, success=True,
                    processing_time_ms=int((time.monotonic() - start) * 1000),
                )
            except Exception as exc:
                return FileProcessingResult(
                    file_path=str(path), file_name=path.name, file_size=size,
                    mime_type=mime, category=self.category, content="",
                    success=False, error=str(exc),
                    processing_time_ms=int((time.monotonic() - start) * 1000),
                )

        # Try pydub for non-wav formats
        try:
            from pydub import AudioSegment
            audio = AudioSegment.from_file(str(path))
            duration = len(audio) / 1000.0
            metadata = {
                "channels": audio.channels,
                "sample_width": audio.sample_width,
                "framerate": audio.frame_rate,
                "duration_seconds": round(duration, 2),
            }
            content = (
                f"[Audio: {path.name}]\n"
                f"Format: {ext.upper().strip('.')}\n"
                f"Duration: {duration:.1f}s\n"
                f"Channels: {audio.channels}\n"
                f"Sample rate: {audio.frame_rate}Hz"
            )
            if progress_callback:
                progress_callback(100)
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category,
                content=content, metadata=metadata, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )
        except ImportError:
            # pydub not installed -- metadata only
            content = f"[Audio: {path.name} ({size:,} bytes, {ext}) -- install pydub for metadata extraction]"
            if progress_callback:
                progress_callback(100)
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category,
                content=content, metadata={}, success=True,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )


class ArchiveHandler(FileTypeHandler):
    """Handles archives via stdlib: .zip, .tar, .gz, .tar.gz"""

    @property
    def extensions(self) -> List[str]:
        return [".zip", ".tar", ".gz", ".tar.gz", ".tgz"]

    @property
    def category(self) -> str:
        return "archive"

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size
        mime = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        metadata: Dict[str, Any] = {}

        name_lower = path.name.lower()
        entries: List[str] = []

        try:
            if name_lower.endswith(".zip"):
                with zipfile.ZipFile(path, "r") as zf:
                    info_list = zf.infolist()
                    metadata["entry_count"] = len(info_list)
                    total_uncompressed = sum(i.file_size for i in info_list)
                    metadata["total_uncompressed"] = total_uncompressed
                    for info in info_list[:200]:
                        entries.append(f"  {info.filename} ({info.file_size:,} bytes)")
                    if len(info_list) > 200:
                        entries.append(f"  ... and {len(info_list) - 200} more entries")

            elif name_lower.endswith((".tar", ".tar.gz", ".tgz", ".gz")):
                mode = "r:gz" if name_lower.endswith((".tar.gz", ".tgz", ".gz")) else "r"
                try:
                    with tarfile.open(path, mode) as tf:
                        members = tf.getmembers()
                        metadata["entry_count"] = len(members)
                        for m in members[:200]:
                            entries.append(f"  {m.name} ({m.size:,} bytes)")
                        if len(members) > 200:
                            entries.append(f"  ... and {len(members) - 200} more entries")
                except tarfile.ReadError:
                    # Possibly a plain .gz file, not a tar archive
                    entries.append(f"  [compressed file: {path.name}]")
                    metadata["entry_count"] = 1

            content = f"[Archive: {path.name}]\nEntries ({metadata.get('entry_count', '?')}):\n"
            content += "\n".join(entries)

            if progress_callback:
                progress_callback(100)
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type=mime, category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type=mime, category=self.category,
            content=content[:MAX_CONTENT_LENGTH],
            metadata=metadata, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )


class RTFHandler(FileTypeHandler):
    """Handles .rtf files via striprtf (optional), regex fallback."""

    @property
    def extensions(self) -> List[str]:
        return [".rtf"]

    @property
    def category(self) -> str:
        return "document"

    @property
    def requires(self) -> List[str]:
        return ["striprtf"]

    def is_available(self) -> bool:
        try:
            from striprtf.striprtf import rtf_to_text  # noqa: F401
            return True
        except ImportError:
            return False

    def process(self, path: Path, progress_callback=None) -> FileProcessingResult:
        start = time.monotonic()
        size = path.stat().st_size

        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            return FileProcessingResult(
                file_path=str(path), file_name=path.name, file_size=size,
                mime_type="application/rtf", category=self.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        if self.is_available():
            try:
                from striprtf.striprtf import rtf_to_text
                content = rtf_to_text(raw)
            except Exception:
                content = self._regex_strip(raw)
        else:
            content = self._regex_strip(raw)

        if progress_callback:
            progress_callback(100)

        return FileProcessingResult(
            file_path=str(path), file_name=path.name, file_size=size,
            mime_type="application/rtf", category=self.category,
            content=content[:MAX_CONTENT_LENGTH],
            metadata={"char_count": len(content)}, success=True,
            processing_time_ms=int((time.monotonic() - start) * 1000),
        )

    @staticmethod
    def _regex_strip(rtf_text: str) -> str:
        """Best-effort RTF stripping via regex (lossy but better than raw RTF)."""
        # Remove RTF control groups
        text = re.sub(r"\{\\[^{}]*\}", "", rtf_text)
        # Remove RTF control words
        text = re.sub(r"\\[a-z]+\d*\s?", "", text)
        # Remove remaining braces
        text = re.sub(r"[{}]", "", text)
        return text.strip()


# ===========================================================================
# FileProcessor -- main orchestrator
# ===========================================================================

class FileProcessor:
    """Orchestrates file validation and content extraction.

    Parameters
    ----------
    allowed_roots
        Directories from which files may be read.  Defaults to
        ~/.kait, ~/Documents, ~/Desktop, ~/Downloads.
    """

    def __init__(self, allowed_roots: Optional[List[Path]] = None):
        self._allowed_roots = [
            r.expanduser().resolve() for r in (allowed_roots or _DEFAULT_ALLOWED_ROOTS)
        ]

        # Build handler registry: extension -> handler instance
        # OCR handler takes priority over plain image if available
        self._handlers: Dict[str, FileTypeHandler] = {}
        handler_classes: List[type] = [
            PlainTextHandler,
            CodeFileHandler,
            PDFHandler,
            DocxHandler,
            SpreadsheetHandler,
            ImageOCRHandler,  # registered first so it overrides ImageHandler
            ImageHandler,
            VideoHandler,
            AudioHandler,
            ArchiveHandler,
            RTFHandler,
        ]
        for cls in handler_classes:
            handler = cls()
            for ext in handler.extensions:
                if ext not in self._handlers:
                    self._handlers[ext] = handler

        log.info(
            "FileProcessor ready: %d extensions, %d handlers",
            len(self._handlers),
            len(handler_classes),
        )

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def validate_file(self, path: str) -> Tuple[bool, str]:
        """Check if a file is acceptable for processing.

        Returns (ok, reason).
        """
        p = Path(path).expanduser().resolve()

        if not p.exists():
            return False, f"File not found: {p}"
        if not p.is_file():
            return False, f"Not a file: {p}"

        # Sandbox check
        in_sandbox = False
        for root in self._allowed_roots:
            try:
                p.relative_to(root)
                in_sandbox = True
                break
            except ValueError:
                continue
        if not in_sandbox:
            return False, f"File outside allowed directories: {p}"

        # Size check
        size = p.stat().st_size
        if size > MAX_FILE_SIZE:
            return False, f"File too large: {size:,} bytes (max {MAX_FILE_SIZE:,})"
        if size == 0:
            return False, "File is empty"

        # Extension check
        ext = p.suffix.lower()
        # Also check compound extension (.tar.gz)
        if ext not in self._handlers:
            compound = "".join(p.suffixes[-2:]).lower() if len(p.suffixes) >= 2 else ""
            if compound not in self._handlers:
                return False, f"Unsupported file type: {ext}"

        return True, "ok"

    def get_supported_extensions(self) -> List[str]:
        """Return sorted list of all supported file extensions."""
        return sorted(self._handlers.keys())

    def get_handler(self, path: Path) -> Optional[FileTypeHandler]:
        """Get the appropriate handler for a file path."""
        ext = path.suffix.lower()
        handler = self._handlers.get(ext)
        if handler is None and len(path.suffixes) >= 2:
            compound = "".join(path.suffixes[-2:]).lower()
            handler = self._handlers.get(compound)
        return handler

    # ------------------------------------------------------------------
    # Processing
    # ------------------------------------------------------------------

    def process_file(
        self,
        path: str,
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> FileProcessingResult:
        """Process a single file and return its extracted content."""
        p = Path(path).expanduser().resolve()
        start = time.monotonic()

        # Validate first
        ok, reason = self.validate_file(str(p))
        if not ok:
            return FileProcessingResult(
                file_path=str(p), file_name=p.name if p.exists() else Path(path).name,
                file_size=p.stat().st_size if p.exists() else 0,
                mime_type="unknown", category="unknown", content="",
                success=False, error=reason,
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        handler = self.get_handler(p)
        if handler is None:
            return FileProcessingResult(
                file_path=str(p), file_name=p.name, file_size=p.stat().st_size,
                mime_type="unknown", category="unknown", content="",
                success=False, error=f"No handler for extension: {p.suffix}",
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )

        try:
            result = handler.process(p, progress_callback)
            log.info(
                "Processed %s (%s, %d bytes) in %dms",
                p.name, handler.category, p.stat().st_size, result.processing_time_ms,
            )
            return result
        except Exception as exc:
            log.error("Error processing %s: %s", p.name, exc)
            return FileProcessingResult(
                file_path=str(p), file_name=p.name, file_size=p.stat().st_size,
                mime_type="unknown", category=handler.category, content="",
                success=False, error=str(exc),
                processing_time_ms=int((time.monotonic() - start) * 1000),
            )


# ===========================================================================
# LLM formatting
# ===========================================================================

def format_for_llm(result: FileProcessingResult, max_chars: int = 50_000) -> str:
    """Format a FileProcessingResult as context text for the LLM.

    Produces a structured block like::

        [Attached File: report.pdf (2.4 MB, document)]
        Metadata: page_count=12, author=...
        Content:
        ...extracted text...
    """
    size_str = _human_size(result.file_size)
    header = f"[Attached File: {result.file_name} ({size_str}, {result.category})]"

    parts = [header]

    if result.metadata:
        meta_items = []
        for k, v in result.metadata.items():
            if isinstance(v, dict):
                continue  # skip nested dicts like exif
            meta_items.append(f"{k}={v}")
        if meta_items:
            parts.append("Metadata: " + ", ".join(meta_items[:15]))

    if not result.success:
        parts.append(f"Error: {result.error}")
    elif result.content:
        content = result.content[:max_chars]
        if len(result.content) > max_chars:
            content += f"\n... [truncated at {max_chars:,} chars]"
        parts.append(f"Content:\n{content}")

    return "\n".join(parts)


def _human_size(size_bytes: int) -> str:
    """Convert byte count to human-readable string."""
    for unit in ("B", "KB", "MB", "GB"):
        if abs(size_bytes) < 1024:
            return f"{size_bytes:.1f} {unit}" if unit != "B" else f"{size_bytes} {unit}"
        size_bytes /= 1024  # type: ignore[assignment]
    return f"{size_bytes:.1f} TB"
